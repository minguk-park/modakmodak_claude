#!/usr/bin/env python3
"""
기능 요구사항 명세서 엑셀 생성 스크립트

사용법:
    python create_excel.py --output 명세서.xlsx --data data.json
    python create_excel.py --output 명세서.xlsx --sample
"""

import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    exit(1)


# 스타일 정의
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def apply_header_style(cell):
    """헤더 셀 스타일 적용"""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def apply_cell_style(cell):
    """일반 셀 스타일 적용"""
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER


def create_requirement_excel(data, output_path):
    """요구사항 명세서 엑셀 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "기능명세서"

    # 헤더 설정
    headers = ["번호", "카테고리", "기능/요구사항명", "상세설명", "비고"]
    column_widths = [8, 15, 25, 60, 20]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = column_widths[col - 1]

    # 데이터 입력
    requirements = data.get("requirements", [])
    for row_idx, req in enumerate(requirements, start=2):
        values = [
            req.get("no", row_idx - 1),
            req.get("category", ""),
            req.get("name", ""),
            req.get("description", ""),
            req.get("note", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            apply_cell_style(cell)

    # 행 높이 자동 조정 (상세설명이 긴 경우)
    for row in range(2, len(requirements) + 2):
        desc = ws.cell(row=row, column=4).value
        if desc:
            line_count = desc.count('\n') + 1
            ws.row_dimensions[row].height = max(30, line_count * 18)

    # 같은 카테고리 셀 병합
    if requirements:
        merge_start = 2
        current_category = requirements[0].get("category", "")

        for row_idx, req in enumerate(requirements[1:], start=3):
            category = req.get("category", "")
            if category != current_category:
                # 이전 카테고리 병합
                if merge_start < row_idx - 1:
                    ws.merge_cells(start_row=merge_start, start_column=2, end_row=row_idx - 1, end_column=2)
                    # 병합된 셀 가운데 정렬
                    ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                merge_start = row_idx
                current_category = category

        # 마지막 카테고리 병합
        last_row = len(requirements) + 1
        if merge_start < last_row:
            ws.merge_cells(start_row=merge_start, start_column=2, end_row=last_row, end_column=2)
            ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(output_path)
    print(f"엑셀 파일 생성 완료: {output_path}")


def get_sample_data():
    """샘플 데이터 반환"""
    return {
        "requirements": [
            {
                "no": 1,
                "category": "인증",
                "name": "회원가입",
                "description": "1. 이메일, 이름, 나이, 휴대전화 번호를 입력받는다\n2. 이메일은 중복되지 않는다.",
                "note": ""
            },
            {
                "no": 2,
                "category": "인증",
                "name": "로그인",
                "description": "1. 이메일과 비밀번호로 로그인한다\n2. 로그인 실패 시 에러 메시지를 표시한다",
                "note": ""
            },
            {
                "no": 3,
                "category": "인증",
                "name": "로그아웃",
                "description": "1. 로그아웃 버튼 클릭 시 세션을 종료한다\n2. 로그인 페이지로 이동한다",
                "note": ""
            }
        ]
    }


def main():
    parser = argparse.ArgumentParser(description="기능 요구사항 명세서 엑셀 생성")
    parser.add_argument("--output", "-o", default="기능명세서.xlsx", help="출력 파일 경로")
    parser.add_argument("--data", "-d", help="JSON 데이터 파일 경로")
    parser.add_argument("--sample", action="store_true", help="샘플 데이터로 생성")
    args = parser.parse_args()

    if args.sample:
        data = get_sample_data()
    elif args.data:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        print("JSON 데이터를 입력하세요 (Ctrl+D로 종료):")
        import sys
        data = json.load(sys.stdin)

    create_requirement_excel(data, args.output)


if __name__ == "__main__":
    main()
